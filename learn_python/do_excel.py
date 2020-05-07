#coding: utf-8
#author = hewangtong
#date = 2020/5/1
#函数一：读取指定表表单数据，并并且把每个单元格数据存储到列表里面，每一行的数据存储到列表里面
#函数二：在指定的单元格写入指定的数据，并保存到当前单元格内
#函数三：新建一个excel
from openpyxl import workbook
from openpyxl import load_workbook
import configparser

cf = configparser.ConfigParser()
cf.read('lemon.conf',encoding='utf-8')
button = cf.get('TestCase','button')
print(button)

class DoExcel:
    '''类的作用是完成excel数据的读写和新建表单的操作'''
    def read_excel(self, file_name, sheet_name):
        '''读取所有的数据，以嵌套列表的形式存储，每一行都是一个子列表，每一个单元格都是子列表中的元素'''
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        if button =='1':
            test_data = []  # 存储所有数据
            for i in range(2, sheet.max_row+1):
                row_data = []   #存储每一行数据
                for j in range(1, sheet.max_column+1):
                    row_data.append(sheet.cell(i, j).value)
                test_data.append(row_data)
        else:
            test_data = []  # 存储所有数据
            for i in eval(button):
                row_data = []  # 存储每一行数据
                for j in range(1, sheet.max_column + 1):
                    row_data.append(sheet.cell(i+1, j).value)
                test_data.append(row_data)

        # '''读取所有的数据，以嵌套字典的形式存储，每一行都是一个字典'''
        # test_data = []  # 存储所有数据
        # for i in range(2, sheet.max_row + 1):
        #     row_data = {}
        #     row_data['Case_id'] = sheet.cell(i, 1).value
        #     row_data['Title'] = sheet.cell(i, 2).value
        #     row_data['Module'] = sheet.cell(i, 3).value
        #     row_data['TestData'] = sheet.cell(i, 4).value
        #     row_data['Expect'] = sheet.cell(i, 5).value
        #     row_data['Result'] = sheet.cell(i, 6).value
        #     row_data['TestResult'] = sheet.cell(i, 7).value
        #     test_data.append(row_data)
        return test_data

    def write_excel(self, file_name, sheet_name,row, col, value):
        '''在指定的单元格写入指定的数据，并保存到当前单元格内'''
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        value = sheet.cell(row, col).value
        wb.save(file_name)
        wb.close()   #每次操作完关闭掉

    def create_excel(self, file_name, sheet_name):
        '''新建一个excel'''
        wb = workbook.Workbook()
        wb.create_sheet(sheet_name)
        wb.save(file_name)

if __name__ == '__main__':
    res = DoExcel().read_excel('python14.xlsx','testcase')
    print(res)
