__author__ = '何旺彤'
from openpyxl import load_workbook
from complete_test_api_test.common.read_config import ReadConfig
from complete_test_api_test.common import project_path


class DoExcel:
    '''该类完成测试数据的读取以及测试结果的写入'''
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name#工作簿文件名地址
        self.sheet_name=sheet_name#表单名

    def read_data(self,section):#配置文件里面的片段名，可以根据你的指定执行指定的用例
        '''从Excel读取数据，有返回值'''
        #从配置文件里面读取哪些测试用例
        case_id=ReadConfig(project_path.conf_path).get_data(section,'case_id')
        wb= load_workbook(self.file_name)#打开工作簿
        sheet=wb[self.sheet_name]#确认表单
        #唯一的要求是什么？每一行数据要在一起{}[]
        #如何把每一行的数据存到一个空间里面去[]
        #开始读取数据
        #获取存在excel里面的电话号码
        tel=self.get_tel()

        test_data=[]
        for i in range(2,sheet.max_row+1):
            row_data={}
            row_data['Caseid']=sheet.cell(i,1).value
            row_data['Module']=sheet.cell(i,2).value
            row_data['Title']=sheet.cell(i,3).value
            row_data['URL']=sheet.cell(i,4).value
            row_data['Method']=sheet.cell(i,5).value
            if sheet.cell(i,6).value.find('tel')!=-1:#-1表示没有找到，也可以用成员运算符
                row_data['Params']=sheet.cell(i,6).value.replace('tel',str(tel))
                self.update_tel(int(tel)+1)
            else:#如果没有找到，不需要做替换
                row_data['Params']=sheet.cell(i,6).value
            row_data['sql']=sheet.cell(i,7).value
            row_data['ExpectedResult']=sheet.cell(i,8).value

            test_data.append(row_data)
        wb.close()
        final_data=[]#空列表存储最终的测试用例
        if case_id=='all': #如果case_id==all那就执行所有用例，否则，如果是列表那就获取列表中的数据
            test_data=test_data
        else:#否则如果是列表，那就获取列表里面指定id用例的测试数据
            for i in case_id:#遍历列表中
                final_data.append(test_data[i-1])
        return test_data

    def get_tel(self):
        '''获取excel表格里面的手机号码'''
        wb= load_workbook(self.file_name)#打开工作簿
        sheet=wb[self.sheet_name]#确认表单
        wb.colse()
        return sheet.cell(1,2).value#返回电话号码的值

    def update_tel(self,new_tel):
        '''写回手机号码'''
        wb=load_workbook(self.file_name)
        sheet=wb['tel']
        sheet.cell(1,2,new_tel)
        wb.save(self.file_name)
        wb.close()

    def write_back(self,row,col,value):
        '''写回测试结果到表格中'''
        wb= load_workbook(self.file_name)#打开工作簿
        sheet=wb[self.sheet_name]#确认表单

        sheet.cell(row,col).value=value

        wb.save(self.file_name)
        wb.close()

if __name__ == '__main__':
    sheet_name='recharge'
    test_data=DoExcel(project_path.case_path,sheet_name).read_data()
    print(test_data)

